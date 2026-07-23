"""
Export service — PDF and Excel report generation.
Uses ReportLab for PDF and openpyxl for Excel.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict

from config.app_config import REPORTS_DIR

logger = logging.getLogger(__name__)


class ExportService:
    """Generates PDF and Excel reports."""
    
    def __init__(self, company_name: str = "Aashu Jewellers"):
        self.company_name = company_name
        os.makedirs(REPORTS_DIR, exist_ok=True)
    
    def export_to_pdf(self, title: str, headers: List[str], data: List[List],
                      filename: str = None, subtitle: str = "") -> str:
        """
        Generate a PDF report with table data.
        
        Args:
            title: Report title
            headers: Column header names
            data: List of row data (list of lists)
            filename: Output filename (auto-generated if None)
            subtitle: Optional subtitle below the title
        
        Returns:
            Path to the generated PDF file
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        except ImportError:
            raise RuntimeError("reportlab not installed. Run: pip install reportlab")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.pdf"
        
        filepath = os.path.join(REPORTS_DIR, filename)
        
        # Determine orientation based on number of columns
        page_size = landscape(A4) if len(headers) > 7 else A4
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=page_size,
            topMargin=20 * mm,
            bottomMargin=15 * mm,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Company header
        company_style = ParagraphStyle(
            "CompanyHeader",
            parent=styles["Title"],
            fontSize=18,
            textColor=colors.HexColor("#1a1d2e"),
            spaceAfter=4,
            alignment=TA_CENTER,
        )
        elements.append(Paragraph(self.company_name, company_style))
        
        # Report title
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading2"],
            fontSize=13,
            textColor=colors.HexColor("#4b5563"),
            spaceAfter=2,
            alignment=TA_CENTER,
        )
        elements.append(Paragraph(title, title_style))
        
        # Subtitle
        if subtitle:
            sub_style = ParagraphStyle(
                "SubTitle",
                parent=styles["Normal"],
                fontSize=10,
                textColor=colors.HexColor("#6b7280"),
                spaceAfter=6,
                alignment=TA_CENTER,
            )
            elements.append(Paragraph(subtitle, sub_style))
        
        # Print date
        date_style = ParagraphStyle(
            "PrintDate",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor("#9ca3af"),
            alignment=TA_RIGHT,
        )
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", date_style
        ))
        elements.append(Spacer(1, 8 * mm))
        
        # Build table
        table_data = [headers] + data
        table = Table(table_data, repeatRows=1)
        
        # Professional table styling
        style = TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1d2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGNMENT", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            
            # Body
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1a1d2e")),
            
            # Alignment
            ("ALIGN", (0, 0), (0, -1), "CENTER"),  # SN column centered
        ])
        
        # Alternating row colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f3f4f6"))
        
        table.setStyle(style)
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        logger.info(f"PDF report generated: {filepath}")
        return filepath
    
    def export_to_excel(self, title: str, headers: List[str], data: List[List],
                        filename: str = None, sheet_name: str = "Report") -> str:
        """
        Generate an Excel report.
        
        Returns:
            Path to the generated .xlsx file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.xlsx"
        
        filepath = os.path.join(REPORTS_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styles
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1D2E", end_color="1A1D2E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        body_font = Font(name="Segoe UI", size=9)
        alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        
        # Company header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=self.company_name).font = Font(
            name="Segoe UI", size=16, bold=True, color="1A1D2E"
        )
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        
        # Title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=title).font = Font(
            name="Segoe UI", size=12, color="4B5563"
        )
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        
        # Headers (row 4)
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.border = thin_border
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = alt_fill
        
        # Auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row in data:
                if col_idx - 1 < len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1])))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_length + 4, 30)
        
        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")
        return filepath
